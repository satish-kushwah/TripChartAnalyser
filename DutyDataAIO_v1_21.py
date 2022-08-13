def myfunction(c,wef):
  import openpyxl,os,time
  from openpyxl.styles import Font
  myfont=Font(bold=True)
  path=os.path.dirname(__file__)+"\\App_data\\"
  generated_data_path=os.path.dirname(__file__)+"\\Generated_DutyData\\"

  #c=int(input('Enter 1 for WEEK, 2 for SAT, 3 for SUN: '))
  #wef=input('Enter wef date in dd.mm.yyyy format: ')
  if c==1:
    from App_data import WeekTripsData,WeekKmData
    d=WeekTripsData.weektrips
    d1=WeekKmData.weekkms
    file=open(path+'WeekSignOnOff.txt','r'); SignOnOff=file.readlines(); file.close()
    excelfile='WeekDutyData_wef_'+wef+'.xlsx'
  elif c==2:
    from App_data import SatTripsData,SatKmData
    d=SatTripsData.sattrips
    d1=SatKmData.satkms
    file=open(path+'SatSignOnOff.txt','r'); SignOnOff=file.readlines(); file.close()
    excelfile='SatDutyData_wef_'+wef+'.xlsx'
  elif c==3:
    from App_data import SunTripsData,SunKmData
    d=SunTripsData.suntrips
    d1=SunKmData.sunkms
    file=open(path+'SunSignOnOff.txt','r'); SignOnOff=file.readlines(); file.close()
    excelfile='SunDutyData_wef_'+wef+'.xlsx'

  wb1 = openpyxl.Workbook() #new workbook
  sheet1 = wb1.active
  sheet1.column_dimensions['C'].width=12
  sheet1.column_dimensions['D'].width=18
  sheet1.column_dimensions['E'].width=18
  currentrow=1
  sheet1.cell(row=currentrow,column=1).value='Developer: satishkushwah50@gmail.com'
  currentrow=3
  for i in range(1,len(d)+1):
    if SignOnOff[i-1].split('\t')[1]!='REST':
      sheet1.cell(row=currentrow,column=1).value='Duty# '+str(i)
      sheet1.cell(row=currentrow,column=1).font=myfont
      sheet1.cell(row=currentrow,column=2).value='SignOn'
      sheet1.cell(row=currentrow,column=2).font=myfont
      sheet1.cell(row=currentrow,column=3).value=SignOnOff[i-1].split('\t')[1]+'/'+SignOnOff[i-1].split('\t')[2]
      sheet1.cell(row=currentrow,column=3).font=myfont
      sheet1.cell(row=currentrow,column=4).value='SignOff'
      sheet1.cell(row=currentrow,column=4).font=myfont
      sheet1.cell(row=currentrow,column=5).value=SignOnOff[i-1].split('\t')[3]+'/'+SignOnOff[i-1].split('\t')[4]
      sheet1.cell(row=currentrow,column=5).font=myfont
      sheet1.cell(row=currentrow,column=6).value=d1[i]+' km'
      sheet1.cell(row=currentrow,column=6).font=myfont
      tripcursor=0; tripno=1
      while(tripcursor<len(d[i])):     
        if d[i][tripcursor][0] in ['Board  ','Departure']:
          currentrow+=1
          if d[i][tripcursor][0]=='Board  ':
            sheet1.cell(row=currentrow,column=2).value='Trip '+str(tripno); tripno+=1
          sheet1.cell(row=currentrow,column=3).value=d[i][tripcursor][2] #train no
          sheet1.cell(row=currentrow,column=4).value=d[i][tripcursor][1]
          tripcursor+=1
        elif d[i][tripcursor][0] in ['Deboard','Arrival']:
          if d[i][tripcursor-1][0] in ['Deboard','Arrival']:
            currentrow+=1
          sheet1.cell(row=currentrow,column=5).value=d[i][tripcursor][1] 
          tripcursor+=1
      currentrow+=2

  wb1.save(generated_data_path+excelfile)
  #input('Data created, press enter to exit')

if __name__ == "__main__":
  c=int(input('Enter 1 for WEEK, 2 for SAT, 3 for SUN: '))
  wef=input('Enter wef date in dd.mm.yyyy format: ')
  myfunction(c,wef)
  input('Data created, press enter to exit')