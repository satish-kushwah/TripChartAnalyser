import os
temp_path=os.path.dirname(__file__)+"\\temp\\"
tripchart_path=os.path.dirname(__file__)+"\\tripchart_input\\"
def myfunction(mychoice,totalduties):
	# edited on 12/6/19, 1153
	#create database by reading data from excel
	# change data as per excel sheet and tt
	firstrow=4; trainNoCol=1
	ODcol=7;  OUcol=19;  JDcol=9;  JUcol=17
	InductLocationCol=4;  InductTimeCol=2;  InductDutyCol=3
	StableLocationCol=25; StableTimeCol=23; StableDutyCol=22

	#uncomment below line if stepping back at both ends
	#BCGNDncol=13; JPWUpCol=23 

	import openpyxl,time,pprint
	#mychoice=int(input('Enter 1 for WEEK, 2 for SAT or 3 for SUN: '))
	if mychoice==1:
		f=open(temp_path+'WeekTripsData.py','w') # database name according to tt
		dictname='weektrips' #dictionary name
		tt=openpyxl.load_workbook(tripchart_path+'WEEK.xlsx') # excel file name
		#ttsheet=tt['TC'] # sheet name
	elif mychoice==2:
		f=open(temp_path+'SatTripsData.py','w') # database name according to tt
		dictname='sattrips' #dictionary name
		tt=openpyxl.load_workbook(tripchart_path+'SAT.xlsx') # excel file name
		#ttsheet=tt['TC'] # sheet name
	elif mychoice==3:	
		f=open(temp_path+'SunTripsData.py','w') # database name according to tt
		dictname='suntrips' #dictionary name
		tt=openpyxl.load_workbook(tripchart_path+'SUN.xlsx') # excel file name
		#ttsheet=tt['TC'] # sheet name
	ttsheet=tt.active
	#totalduties=int(input('Enter total no of duties: '))
	totalrows=ttsheet.max_row
	print('Reading',totalrows,'rows')
	allduties={}
	for k in range(1,totalduties+1): 
		duty=k; trainStable=0; trainInduct=0
		alltrips=[]; trip=[]
		for i in range(totalrows): #reading rows of excels
			#boarding condition at down platform
			if ttsheet.cell(row=i+firstrow,column=ODcol-1).value!=None and ttsheet.cell(row=i+firstrow,column=ODcol-1).value!=duty and ttsheet.cell(row=i+firstrow,column=ODcol+1).value==duty:
				trainTime=ttsheet.cell(row=i+firstrow,column=ODcol).value
				if trainTime==None:
					trainTime='Unknown'
				trip.append('Board  '); trip.append(str(trainTime)[-8:-3]+'/DOWN'); trip.append(str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value))
				alltrips.append(trip)
				trip=[]
			#deboarding condition at down platform
			if ttsheet.cell(row=i+firstrow,column=ODcol-1).value==duty and ttsheet.cell(row=i+firstrow,column=ODcol+1).value!=None and ttsheet.cell(row=i+firstrow,column=ODcol+1).value!=duty:
				trainTime=ttsheet.cell(row=i+firstrow,column=ODcol).value
				if trainTime==None:
					trainTime='Unknown'
				trip.append('Deboard'); trip.append(str(trainTime)[-8:-3]+'/DOWN'); trip.append(str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value))
				alltrips.append(trip)
				trip=[]
			#boarding condition at up platform
			if ttsheet.cell(row=i+firstrow,column=OUcol-1).value!=None and ttsheet.cell(row=i+firstrow,column=OUcol-1).value!=duty and ttsheet.cell(row=i+firstrow,column=OUcol+1).value==duty:
				trainTime=ttsheet.cell(row=i+firstrow,column=OUcol).value
				if trainTime==None:
					trainTime='Unknown'
				trip.append('Board  '); trip.append(str(trainTime)[-8:-3]+'/UP  '); trip.append(str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value))
				alltrips.append(trip)
				trip=[]
			#deboarding condition at up platform
			if ttsheet.cell(row=i+firstrow,column=OUcol-1).value==duty and ttsheet.cell(row=i+firstrow,column=OUcol+1).value!=None and ttsheet.cell(row=i+firstrow,column=OUcol+1).value!=duty:
				trainTime=ttsheet.cell(row=i+firstrow,column=OUcol).value
				if trainTime==None:
					trainTime='Unknown'
				trip.append('Deboard'); trip.append(str(trainTime)[-8:-3]+'/UP  '); trip.append(str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value))
				alltrips.append(trip)
				trip=[]
			
			#Boarding condition on train induction
			if ttsheet.cell(row=i+firstrow,column=InductTimeCol).value!=None and ttsheet.cell(row=i+firstrow,column=InductDutyCol).value==duty:
			#if ttsheet.cell(row=i+firstrow,column=InductLocationCol).value!=None and ttsheet.cell(row=i+firstrow,column=InductDutyCol).value==duty:
				trainTime=ttsheet.cell(row=i+firstrow,column=InductTimeCol).value
				if trainTime==None:
					trainTime='Unknown'
				trip.append('Board  '); trip.append(str(trainTime)[-8:-3]+'/'+ttsheet.cell(row=i+firstrow,column=InductLocationCol).value.strip()); trip.append(str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value))
				if ttsheet.cell(row=i+firstrow,column=InductLocationCol).value.strip()=='KKD':
					#checking whether train inducts from JU or JD
					if ttsheet.cell(row=i+firstrow,column=JDcol+1).value!=None: 
						trip[1]=str(trainTime)[-8:-3]+'/KKD_JD' 
					else:
						trip[1]=str(trainTime)[-8:-3]+'/KKD_JU' 
				firstTrip=trip; trainInduct=1
				trip=[]
			#Deborad condition other than OKNS
			if ttsheet.cell(row=i+firstrow,column=StableTimeCol).value!=None and ttsheet.cell(row=i+firstrow,column=StableDutyCol).value==duty:
			#if ttsheet.cell(row=i+firstrow,column=StableLocationCol).value!=None and ttsheet.cell(row=i+firstrow,column=StableDutyCol).value==duty:
				trainTime=ttsheet.cell(row=i+firstrow,column=StableTimeCol).value
				if trainTime==None:
					trainTime='Unknown'
				trip.append('Deboard'); trip.append(str(trainTime)[-8:-3]+'/'+ttsheet.cell(row=i+firstrow,column=StableLocationCol).value.strip()); trip.append(str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value))
				if ttsheet.cell(row=i+firstrow,column=StableLocationCol).value.strip()=="KKD":
					#checking whether train terminates at JU or JD
					if ttsheet.cell(row=i+firstrow,column=JUcol+1).value!=None: 
						trip[1]=str(trainTime)[-8:-3]+'/KKD_JU'; 
					else:
						trip[1]=str(trainTime)[-8:-3]+'/KKD_JD'; 
				lastTrip=trip; trainStable=1
				trip=[]

		for i in range(len(alltrips)):
			for j in range(len(alltrips)-1-i):
				if alltrips[j][1].split('/')[0]>alltrips[j+1][1].split('/')[0]:
					alltrips[j],alltrips[j+1]=alltrips[j+1],alltrips[j]
		if trainStable==1:
			alltrips.append(lastTrip)
		if trainInduct==1:
			alltrips.insert(0,firstTrip)
		allduties[k]=alltrips
	"""
	# Uncomment this section when stepping back at both end 
	#Train change at BCGN and JPW
	for i in range(totalrows):
		#Train change at BCGN
		if ttsheet.cell(row=i+firstrow,column=BCGNDncol-1).value!=None and ttsheet.cell(row=i+firstrow,column=BCGNDncol+5).value!=None and ttsheet.cell(row=i+firstrow,column=BCGNDncol-1).value!=ttsheet.cell(row=i+firstrow,column=BCGNDncol+5).value:
			dutyno=ttsheet.cell(row=i+firstrow,column=BCGNDncol+5).value; trainno=str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value)
			BCGNtime=str(ttsheet.cell(row=i+firstrow,column=BCGNDncol).value); 
			trips=allduties[dutyno]
			for j in range(int(len(trips)/2)):
				if (trips[j*2][1].split('/')[0]<BCGNtime<trips[j*2+1][1].split('/')[0]) and trainno!=trips[j*2][2].split(',')[0] and trainno!=trips[j*2+1][2].split(',')[0]:
					trips[j*2][2]=trips[j*2][2]+',  '+trainno
					allduties[dutyno]=trips
		#Train change at JPW
		if ttsheet.cell(row=i+firstrow,column=JPWUpCol-1).value!=None and ttsheet.cell(row=i+firstrow,column=JPWUpCol+1).value!=None and ttsheet.cell(row=i+firstrow,column=JPWUpCol-1).value!=ttsheet.cell(row=i+firstrow,column=JPWUpCol+1).value:
			dutyno=ttsheet.cell(row=i+firstrow,column=JPWUpCol+1).value; trainno=str(ttsheet.cell(row=i+firstrow,column=trainNoCol).value)
			JPWtime=str(ttsheet.cell(row=i+firstrow,column=JPWUpCol).value); 
			trips=allduties[dutyno]
			for j in range(int(len(trips)/2)):
				if (trips[j*2][1].split('/')[0]<JPWtime<trips[j*2+1][1].split('/')[0]) and trainno!=trips[j*2][2].split(',')[0] and trainno!=trips[j*2+1][2].split(',')[0]:
					trips[j*2][2]=trips[j*2][2]+',  '+trainno
					allduties[dutyno]=trips		
	"""
	f.write(dictname+'='+pprint.pformat(allduties))	
	f.close()
	#input('TripChart created successfully, press enter to exit')

if __name__ == "__main__":
	mychoice=int(input('Enter 1 for WEEK, 2 for SAT or 3 for SUN: '))
	totalduties=int(input('Enter total no of duties: '))
	myfunction(mychoice,totalduties)
	input('TripChart created successfully, press enter to exit')